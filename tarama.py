"""
BIST ESv4 Strateji Tarama — Günlük Periyot
────────────────────────────────────────────
Mantık: Akşam seans kapandıktan sonra çalıştır.
        AL sinyali veren hisseler ertesi gün için listelenir.
        Mum kapanışı beklenir (barstate.isconfirmed mantığı).

Kurulum:
  pip install yfinance pandas numpy openpyxl
Çalıştırma:
  python bist_tarama.py
"""

import pandas as pd
import numpy as np
from datetime import datetime, date
import warnings
import logging
import os
import sys
from pathlib import Path
warnings.filterwarnings("ignore")
logging.getLogger("yfinance").setLevel(logging.CRITICAL)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import yfinance as yf
except ImportError:
    yf = None

try:
    from tvDatafeed import TvDatafeed, Interval
except ImportError:
    TvDatafeed = None
    Interval = None

SCRIPT_VERSION = "tarama.py 2026-06-08 ESv4-revised"

# ─────────────────────────────────────────────
# PARAMETRELER
# ─────────────────────────────────────────────
INTERVAL    = "1d"     # Günlük periyot
PERIOD_1D   = "2y"     # EMA/MA ve pozisyon akisi icin yeterli gecmis
PERIOD_HTF  = "2y"
DATA_SOURCE = "tradingview"  # Ana veri kaynagi: yfinance veya tradingview
ALLOW_DATA_FALLBACK = False  # Tutarlilik icin tek kaynak kullan
TV_EXCHANGE = "BIST"
TV_RETRY_COUNT = 3
TV_PREFLIGHT_SYMBOL = "THYAO"
LAST_DATA_SOURCE_ERROR = ""
MED_LEN     = 3
RSI_LEN     = 14
STOCH_LEN   = 14
SMOOTH_K    = 3
SMOOTH_D    = 3
EMA_LEN     = 14
LOOKBACK    = 4
VOL_LEN     = 20
HTF_MA_LEN  = 200
STRONG_MAX  = 30.0
WEAK_MIN    = 70.0
GRADE_STRONG_MAX = min(STRONG_MAX, WEAK_MIN)
GRADE_WEAK_MIN = max(STRONG_MAX, WEAK_MIN)
STOP_LOSS_PCT = 5.0
PROFIT_TRIGGER_PCT = 15.0
PULLBACK_PCT = 5.0
MA_TREND_LEN = 20
MA_SLOPE_BARS = 5
MIN_MA_SLOPE_PCT = 0.5
SIGNAL_VOLUME_MULTIPLIER = 1
USE_SIGNAL_VOLUME_FILTER = True # Hacim aç-kapa
USE_HTF     = False    # Varsayılan OFF
USE_TREND   = True     # MA20 slope-only filtresi
TAKIP_EXCEL_DOSYASI = "Yüzde .xlsx"
TAKIP_EXCEL_SAYFA = "VERİLER"
TAKIP_EXCEL_GUNCELLE = True
GUNLUK_AL_EXCEL_KAYDET = False
TAKIP_KOLONLARI = [
    "İşlem ID",
    "Hisse",
    "Sektör",
    "AL Tarihi",
    "AL Gücü",
    "Giriş Fiyatı",
    "Stop Fiyatı",
    "Durum",
    "Çıkış Tarihi",
    "Çıkış Fiyatı",
    "Çıkış Nedeni",
    "Güncel Fiyat",
    "Gerçekleşen Getiri (%)",
    "Güncel Getiri (%)",
    "Pozisyonda Gün",
    "En Yüksek Fiyat",
    "En Düşük Fiyat",
    "MFE (%)",
    "MAE (%)",
    "BIST Giriş",
    "BIST Çıkış",
    "BIST Getiri (%)",
    "Alpha (%)",
    "Stop Sonrası Güncel (%)",
    "Not",
]
# ─────────────────────────────────────────────
# BIST HİSSE LİSTESİ
# ─────────────────────────────────────────────
BIST_HISSELER = sorted(list(set([
    "ACSEL","ADEL","AEFES","AFYON","AGESA","AGHOL","AKBNK","AKCNS","AKFYE",
    "AKGRT","AKMGY","AKSA","AKSEN","AKSGY","ALARK","ALBRK","ALFAS","ALKIM",
    "ALMAD","ALVARK","ANELE","ARCLK","ARDYZ","ARENA","ARSAN","ASELS","ASTOR",
    "ATEKS","AYGAZ","BAGFS","BAKAB","BANVT","BERA","BFREN","BIMAS","BJKAS",
    "BOSSA","BRISA","BRMEN","BRSAN","BTCIM","BUCIM","BURCE","BURVA","CCOLA",
    "CELHA","CEMAS","CEMTS","CIMSA","CLEBI","CONAS","CWENE","DEVA","DITAS",
    "DMSAS","DOAS","DOBUR","DOHOL","DOKTA","DYOBY","ECILC","ECZYT","EGEEN",
    "EGEPO","EGGUB","EGPRO","EKGYO","EREGL","FROTO","GARAN","GENIL","GENTS",
    "GEREL","GESAN","GOLTS","GOODY","GOZDE","GUBRF","HALKB","HATEK","HEKTS",
    "HUBVC","HURGZ","ICBCT","INDES","INFO","INVEO","ISDMR","ISFIN","ISGYO",
    "ISKUR","JANTS","KAREL","KARSN","KARTN","KCHOL","KLKIM","KLMSN","KLSYN",
    "KONYA","KORDS","KOZAA","KOZAL","KRDMA","KRDMB","KRDMD","LOGO","MAALT",
    "MAGEN","MAKIM","MARKA","MAVKG","MEDTR","MEPET","MERCN","MERIT","MERKO",
    "METRO","MIGROS","MNDRS","MOBTL","MPARK","MRSHL","NATEN","NETAS","NTGAZ",
    "NTHOL","NTTUR","NUHCM","ODAS","ORGE","ORMA","OTKAR","OYAKC","PETKM",
    "PETUN","PGSUS","PKART","POLHO","PRKAB","PRKME","QNBFB","RYSAS","SAHOL",
    "SANEL","SANKO","SARKY","SASA","SISE","SKBNK","SKTAS","SOKM","TCELL",
    "THYAO","TKFEN","TOASO","TTKOM","TTRAK","TUPRS","TURGG","VAKBN","VAKKO",
    "VESBE","VESTL","YKBNK","YONGA","ZOREN","AKENR","AKFGY","ANHYT","ARAT",
    "ATATP","AVGYO","AYCES","AYEN","BASGZ","BAYRK","BIENY","BINBN","BIOEN",
    "BIZIM","BLCYT","BNTAS","BRYAT","BVSAN","CANTE","CEOEM","CLKHO","CRFSA",
    "CUSAN","CVKMD","DAGHL","DAGI","DAPGM","DARDL","DENGE","DERHL","DESA",
    "DESPC","DGATE","DGGYO","DGNMO","DNISI","DURDO","DZGYO","EDATA","EDIP",
    "EMKEL","EMNIS","ENPRO","ENRUY","ERSU","ESCAR","ESCOM","ESEN","ETILR",
    "EUREN","EUYO","EVCIL","FADE","FENER","FLAP","FONET","FORMT","FORTE",
    "FZLGY","GARFA","GEDIK","GEDZA","GLBMD","GLCVY","GLYHO","GMTAS","GOKNR",
    "GRNYO","GRSEL","GRTRK","GSDDE","GSDHO","GSRAY","GWIND","GZNMI","HDFGS",
    "HEDEF","HKTM","HLGYO","HPGYO","HRKET","HTTBT","HUNER","IDGYO","IEYHO",
    "IHLGM","IHEVA","IHGZT","IHLAS","IMASM","INTEM","IPEKE","ISGSY","ISKPL",
    "ISMO","ISYAT","ITTFH","IZFAS","IZINV","IZMDC","KAPLM","KATMR","KAYSE",
    "KBORU","KCAER","KENT","KERVN","KERVT","KFEIN","KGYO","KIMMR","KLGYO",
    "KLNMA","KLRHO","KLSER","KMPUR","KNFRT","KOCMT","KOPOL","KRONT","KRPLS",
    "KRSTL","KRTEK","KRVGD","KTLEV","KTSKR","KUTPO","KUVVA","KUYAS","LIDER",
    "LIDFA","LILAK","LKMNH","LMKDC","LRSHO","LUKSK","MACKO","MANAS","MARTI",
    "MEGAP","METUR","MIATK","MMCAS","MNDTR","MOGAN","MRGYO","MSGYO","MTRKS",
    "MZHLD","NIBAS","NUGYO","OBAMS","OBASE","ODINE","OFKGT","ONCSM","ORCAY",
    "OSMEN","OSTIM","OYAYO","OYLUM","OZGYO","OZKGY","OZRDN","OZSUB","PAGYO",
    "PAMEL","PAPIL","PARSN","PASEU","PCILT","PEGYO","PEKMT","PENGD","PENTA",
    "PINSU","PKENT","PLTUR","PNLSN","POLTK","PRDGS","PRZMA","PSDTC","PSGYO",
    "QNBFL","RALYH","RAYSG","RHEAG","RNPOL","RODRG","ROYAL","RTALB","RUBNS",
    "SAYAS","SDTTR","SEGYO","SEKFK","SEKUR","SELEC","SELGD","SELVA","SEYKM",
    "SILVR","SMART","SMRTG","SNGYO","SNICA","SNKRN","SONME","SRVGY","SUMAS",
    "SUNTK","SUWEN","TABGD","TARKM","TATEN","TATGD","TAVHL","TBORG","TDGYO",
    "TEKTU","TERA","TETMT","TGSAS","TKNSA","TLMAN","TMSN","TNZTP","TRCAS",
    "TRGYO","TRILC","TSGYO","TSPOR","TUCLK","TUKAS","ULUUN","ULUSE","UMPAS",
    "UNLU","USAK","VAKFN","VANGD","VBTYZ","VERUS","VKFYO","VKGYO","VKING",
    "YATAS","YAYLA","YBTAS","YEOTK","YGGYO","YKSLN","YOYGD","YPKYO","YUNSA",
    "ZEDUR","ZRGYO"
])))

# ─────────────────────────────────────────────
# FONKSİYONLAR
# ─────────────────────────────────────────────
def percentile_nearest_rank(series, length, pct):
    result = series.copy() * np.nan
    arr    = series.values
    for i in range(length - 1, len(arr)):
        w = arr[i - length + 1:i + 1]
        w = w[~np.isnan(w)]
        if len(w) == 0:
            continue
        idx = int(np.ceil(pct / 100.0 * len(w))) - 1
        result.iloc[i] = np.sort(w)[max(0, min(idx, len(w)-1))]
    return result

def ema(series, length):
    return series.ewm(span=length, adjust=False).mean()

def sma(series, length):
    return series.rolling(window=length).mean()

def rma(series, length):
    values = series.astype(float)
    result = pd.Series(np.nan, index=series.index, dtype=float)
    seed = values.rolling(window=length, min_periods=length).mean()

    for i in range(len(values)):
        value = values.iloc[i]
        if np.isnan(value):
            continue
        previous_is_empty = i == 0 or np.isnan(result.iloc[i - 1])
        if previous_is_empty:
            if not np.isnan(seed.iloc[i]):
                result.iloc[i] = seed.iloc[i]
        else:
            result.iloc[i] = (result.iloc[i - 1] * (length - 1) + value) / length

    return result

def adx_calc(high, low, close, length):
    up_move = high.diff()
    down_move = -low.diff()
    plus_dm = pd.Series(np.where((up_move > down_move) & (up_move > 0), up_move, 0.0), index=high.index)
    minus_dm = pd.Series(np.where((down_move > up_move) & (down_move > 0), down_move, 0.0), index=high.index)

    tr1 = high - low
    tr2 = (high - close.shift(1)).abs()
    tr3 = (low - close.shift(1)).abs()
    true_range = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)

    atr = rma(true_range, length)
    plus_di = 100.0 * rma(plus_dm, length) / atr
    minus_di = 100.0 * rma(minus_dm, length) / atr
    dx = 100.0 * (plus_di - minus_di).abs() / (plus_di + minus_di)
    return rma(dx.replace([np.inf, -np.inf], np.nan), length)

def rsi_calc(close, length):
    delta    = close.diff()
    gain     = delta.clip(lower=0)
    loss     = (-delta).clip(lower=0)
    avg_gain = rma(gain, length)
    avg_loss = rma(loss, length)
    rs       = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def stoch_rsi(close, rsi_len, stoch_len, smooth_k, smooth_d):
    rsi_val   = rsi_calc(close, rsi_len)
    rsi_min   = rsi_val.rolling(stoch_len).min()
    rsi_max   = rsi_val.rolling(stoch_len).max()
    stoch_raw = (rsi_val - rsi_min) / (rsi_max - rsi_min + 1e-10) * 100
    K         = sma(stoch_raw, smooth_k)
    D         = sma(K, smooth_d)
    return K, D

def crossover_win(a, b, n):
    cross = (a > b) & (a.shift(1) <= b.shift(1))
    return cross.rolling(n).max().fillna(0).astype(bool)

def valuewhen(cond, value):
    return value.where(cond).ffill()

def buy_grade_text(grade):
    if grade == 3:
        return "GUCLU AL"
    if grade == 2:
        return "NORMAL AL"
    if grade == 1:
        return "ZAYIF AL"
    return "AL"

TV_CLIENT = None
LOCAL_ENV_LOADED = False

def local_env_yukle():
    global LOCAL_ENV_LOADED
    if LOCAL_ENV_LOADED:
        return
    LOCAL_ENV_LOADED = True

    env_path = Path(".env")
    if not env_path.exists():
        return

    try:
        for raw_line in env_path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value
    except Exception:
        pass

def normalize_ohlcv(df):
    if df is None or len(df) == 0:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = df.columns.get_level_values(0)
    rename_map = {col: str(col).title() for col in df.columns}
    df = df.rename(columns=rename_map)
    required = ["Open", "High", "Low", "Close", "Volume"]
    if not all(col in df.columns for col in required):
        return None
    return df[required].dropna(subset=["High", "Low", "Close"])

def period_to_bars(period):
    if period.endswith("y"):
        return int(period[:-1]) * 260
    if period.endswith("mo"):
        return int(period[:-2]) * 22
    if period.endswith("d"):
        return int(period[:-1])
    return 600

def tv_interval(interval):
    if Interval is None:
        return None
    if interval == "1d":
        return Interval.in_daily
    return None

def tradingview_symbol(ticker):
    return ticker.replace(".IS", "")

def get_tv_client():
    global TV_CLIENT
    if TvDatafeed is None:
        return None
    local_env_yukle()
    if TV_CLIENT is None:
        username = os.getenv("TV_USERNAME")
        password = os.getenv("TV_PASSWORD")
        if username and password:
            TV_CLIENT = TvDatafeed(username=username, password=password)
        else:
            TV_CLIENT = TvDatafeed()
    return TV_CLIENT

def veri_cek_yfinance(ticker, period, interval):
    if yf is None:
        return None
    df = yf.download(ticker, period=period, interval=interval,
                     progress=False, auto_adjust=True)
    return normalize_ohlcv(df)

def veri_cek_tradingview(ticker, period, interval):
    tv_int = tv_interval(interval)
    if TvDatafeed is None or tv_int is None:
        return None
    last_error = None
    for _ in range(TV_RETRY_COUNT):
        try:
            client = get_tv_client()
            if client is None:
                return None
            df = client.get_hist(
                symbol=tradingview_symbol(ticker),
                exchange=TV_EXCHANGE,
                interval=tv_int,
                n_bars=period_to_bars(period)
            )
            normalized = normalize_ohlcv(df)
            if normalized is not None and len(normalized) > 0:
                return normalized
            last_error = "veri yok"
        except Exception as e:
            global TV_CLIENT
            TV_CLIENT = None
            last_error = str(e)
    raise RuntimeError(last_error or "TradingView veri alinamadi")

def veri_cek_kaynakli(ticker, period, interval):
    global LAST_DATA_SOURCE_ERROR
    LAST_DATA_SOURCE_ERROR = ""
    sources = [DATA_SOURCE]
    fallback = "yfinance" if DATA_SOURCE == "tradingview" else "tradingview"
    if ALLOW_DATA_FALLBACK and fallback not in sources:
        sources.append(fallback)

    for source in sources:
        try:
            if source == "tradingview":
                df = veri_cek_tradingview(ticker, period, interval)
            elif source == "yfinance":
                df = veri_cek_yfinance(ticker, period, interval)
            else:
                df = None
            if df is not None and len(df) > 0:
                return df, source
            LAST_DATA_SOURCE_ERROR = f"{source}: veri yok"
        except Exception as e:
            LAST_DATA_SOURCE_ERROR = f"{source}: hata - {e}"
            continue
    return None, ""

def son_veri_kaynagi_hatasi():
    return LAST_DATA_SOURCE_ERROR

def tradingview_on_kontrol():
    if DATA_SOURCE != "tradingview":
        return True, "TradingView ana kaynak degil"
    if TvDatafeed is None:
        return False, (
            "tvDatafeed bu Python ortaminda kurulu degil. "
            f"Python: {sys.executable} | Kurulum: python -m pip install tvdatafeed"
        )

    try:
        test_ticker = TV_PREFLIGHT_SYMBOL + ".IS"
        df = veri_cek_tradingview(test_ticker, "1mo", INTERVAL)
        if df is None or len(df) == 0:
            return False, f"{TV_PREFLIGHT_SYMBOL} icin TradingView veri donmedi"

        login_text = "girisli" if os.getenv("TV_USERNAME") and os.getenv("TV_PASSWORD") else "nologin"
        return True, f"TradingView baglantisi OK ({login_text}) | test: {TV_PREFLIGHT_SYMBOL}, satir: {len(df)}"
    except Exception as e:
        return False, f"TradingView baglantisi basarisiz: {e}"

def veri_cek(ticker, period, interval):
    df, _ = veri_cek_kaynakli(ticker, period, interval)
    return df

def htf_ok(ticker):
    try:
        df = veri_cek(ticker, PERIOD_HTF, "1d")
        if df is None or len(df) < HTF_MA_LEN + 5:
            return True
        close = df["Close"]
        ma    = sma(close, HTF_MA_LEN)
        return bool(float(close.iloc[-1]) > float(ma.iloc[-1]))
    except:
        return True

def sinyal_hesapla(df):
    high  = df["High"]
    low   = df["Low"]
    close = df["Close"]
    vol   = df["Volume"]
    hl2   = (high + low) / 2

    median     = percentile_nearest_rank(hl2, MED_LEN, 50)
    median_ema = ema(median, MED_LEN)

    K, D  = stoch_rsi(close, RSI_LEN, STOCH_LEN, SMOOTH_K, SMOOTH_D)
    ema_k = ema(K, EMA_LEN)

    cross3_raw = (K > ema_k) & (K.shift(1) <= ema_k.shift(1))
    c1 = crossover_win(median, median_ema, LOOKBACK)
    c2 = crossover_win(K, D, LOOKBACK)
    c3 = cross3_raw.rolling(LOOKBACK).max().fillna(0).astype(bool)

    ma_trend = sma(close, MA_TREND_LEN)
    ma_slope_ok = ma_trend >= ma_trend.shift(MA_SLOPE_BARS) * (1.0 + MIN_MA_SLOPE_PCT / 100.0)
    trend_ok = ma_slope_ok if USE_TREND else pd.Series(True, index=close.index)
    vol_avg = sma(vol, VOL_LEN)
    if USE_SIGNAL_VOLUME_FILTER:
        signal_vol_ok = vol >= vol_avg * SIGNAL_VOLUME_MULTIPLIER
    else:
        signal_vol_ok = pd.Series(True, index=close.index)
    setup_repeated = c1.shift(1).fillna(False) & c2.shift(1).fillna(False) & c3.shift(1).fillna(False)
    long_raw = c1 & c2 & c3 & trend_ok & signal_vol_ok & ~setup_repeated
    # sat_raw sadece bilgi amacli hesaplaniyor, pozisyon kapatmiyor
    sat_raw = pd.Series(False, index=close.index)

    cross_level = valuewhen(cross3_raw, K)
    grade = pd.Series(0, index=close.index)
    grade = grade.mask(long_raw & (cross_level <= GRADE_STRONG_MAX), 3)
    grade = grade.mask(long_raw & (cross_level > GRADE_STRONG_MAX) & (cross_level < GRADE_WEAK_MIN), 2)
    grade = grade.mask(long_raw & (cross_level >= GRADE_WEAK_MIN), 1)

    # Pozisyon acikken tekrar AL uretmemek icin gunluk AL/SAT akisini takip et.
    al_sinyal  = pd.Series(False, index=close.index)
    sat_sinyal = pd.Series(False, index=close.index)
    grade_sinyal = pd.Series(0, index=close.index)
    stop_fiyat = pd.Series(np.nan, index=close.index)
    sat_neden = pd.Series("", index=close.index)
    pozisyon_acik = False
    entry_price = np.nan
    active_stop = np.nan
    highest_since_entry = np.nan

    for i in range(len(close)):
        fiyat = float(close.iloc[i])
        low_i = float(low.iloc[i])
        high_i = float(high.iloc[i])

        if not pozisyon_acik and bool(long_raw.iloc[i]):
            al_sinyal.iloc[i] = True
            grade_sinyal.iloc[i] = int(grade.iloc[i])
            stop_fiyat.iloc[i] = fiyat * (1.0 - STOP_LOSS_PCT / 100.0)
            pozisyon_acik = True
            entry_price = fiyat
            active_stop = stop_fiyat.iloc[i]
            highest_since_entry = high_i
            continue

        if pozisyon_acik:
            highest_since_entry = max(highest_since_entry, high_i)
            peak_profit_pct = (highest_since_entry / entry_price - 1.0) * 100.0 if entry_price > 0 else 0.0
            pullback_from_peak_pct = (highest_since_entry / fiyat - 1.0) * 100.0 if fiyat > 0 else 0.0
            stop_raw = low_i <= active_stop
            kar_stop_raw = peak_profit_pct >= PROFIT_TRIGGER_PCT and pullback_from_peak_pct >= PULLBACK_PCT

        if pozisyon_acik and (stop_raw or kar_stop_raw or bool(sat_raw.iloc[i])):
            sat_sinyal.iloc[i] = True
            sat_neden.iloc[i] = "STOP SAT" if stop_raw else "KAR STOP" if kar_stop_raw else "SAT"
            pozisyon_acik = False
            entry_price = np.nan
            active_stop = np.nan
            highest_since_entry = np.nan

    return al_sinyal, sat_sinyal, close, grade_sinyal, stop_fiyat, sat_neden


def gunluk_al_tara(symbols=None, log_func=None):
    """tarama.py ile daily_scan_telegram.py ayni AL sonucunu uretsin."""
    symbols = BIST_HISSELER if symbols is None else symbols
    al_listesi = []
    hata_listesi = []
    toplam = len(symbols)

    def log(message):
        if log_func is not None:
            log_func(message)

    if DATA_SOURCE == "tradingview":
        ok, mesaj = tradingview_on_kontrol()
        log(f"Veri on kontrol: {mesaj}")
        if not ok:
            return al_listesi, list(symbols)

    for idx, hisse in enumerate(symbols, 1):
        ticker = hisse + ".IS"
        log(f"[{idx:3d}/{toplam}] {hisse}: taraniyor")
        try:
            if USE_HTF and not htf_ok(ticker):
                log(f"{hisse}: HTF engelledi")
                continue

            df, veri_kaynagi = veri_cek_kaynakli(ticker, PERIOD_1D, INTERVAL)
            if df is None or len(df) < max(30, MA_TREND_LEN + MA_SLOPE_BARS + 5, VOL_LEN + 5):
                hata = son_veri_kaynagi_hatasi()
                kaynak_text = veri_kaynagi if veri_kaynagi else "yok"
                log(f"{hisse}: veri yok | veri: {kaynak_text}" + (f" | {hata}" if hata else ""))
                hata_listesi.append(hisse)
                continue

            al, sat, close, grade, stop_fiyat, sat_neden = sinyal_hesapla(df)

            if len(al) < 3:
                log(f"{hisse}: yetersiz veri")
                continue

            son_al = bool(al.iloc[-1])
            son_sat = bool(sat.iloc[-1])

            if son_al:
                sinyal_tarihi = df.index[-1].strftime("%d.%m.%Y")
                sinyal_fiyat = round(float(close.iloc[-1]), 2)
                stop_seviye = round(float(stop_fiyat.iloc[-1]), 2)
                al_gucu = buy_grade_text(int(grade.iloc[-1]))

                al_listesi.append({
                    "Hisse": hisse,
                    "Kapanış Fiyatı": sinyal_fiyat,
                    "Stop Fiyatı": stop_seviye,
                    "AL Gücü": al_gucu,
                    "Sinyal Tarihi": sinyal_tarihi,
                    "Veri Kaynagi": veri_kaynagi,
                    "Not": "Ertesi gün açılışta giriş"
                })
                log(f"{hisse}: {al_gucu} sinyali bulundu @ {sinyal_fiyat} stop {stop_seviye} | veri: {veri_kaynagi}")
            elif son_sat:
                neden = sat_neden.iloc[-1] if sat_neden.iloc[-1] else "SAT"
                log(f"{hisse}: {neden} | veri: {veri_kaynagi}")
            else:
                log(f"{hisse}: sinyal yok | veri: {veri_kaynagi}")

        except Exception as e:
            log(f"{hisse}: hata - {e}")
            hata_listesi.append(hisse)

    return al_listesi, hata_listesi

def takip_dosyasi_bul():
    hedef = Path(TAKIP_EXCEL_DOSYASI)
    if hedef.exists():
        return hedef

    for dosya in Path(".").glob("*.xlsx"):
        if dosya.name.startswith("~$"):
            continue
        try:
            if load_workbook is None:
                continue
            wb = load_workbook(dosya, read_only=True, data_only=True)
            if TAKIP_EXCEL_SAYFA in wb.sheetnames:
                wb.close()
                return dosya
            wb.close()
        except Exception:
            continue
    return hedef

def hucre_bos_mu(value):
    if value is None:
        return True
    if isinstance(value, float) and np.isnan(value):
        return True
    return str(value).strip() == ""

def hisse_kodu_temizle(value):
    if hucre_bos_mu(value):
        return ""
    return str(value).strip().upper().replace(".IS", "")

def tarih_degeri(value):
    if hucre_bos_mu(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date()

def baslik_haritasi(ws):
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if not hucre_bos_mu(cell.value)
    }

def hucre_degeri(ws, row, headers, *names):
    for name in names:
        col = headers.get(name)
        if col is not None:
            return ws.cell(row=row, column=col).value
    return None

def yuvarla(value, digits=2):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return round(float(value), digits)

def getiri_yuzde(baslangic, bitis):
    if baslangic is None or bitis is None or float(baslangic) == 0:
        return None
    return round((float(bitis) / float(baslangic) - 1.0) * 100.0, 2)

def takip_islem_id(row):
    return f"ISLEM-{row - 1:06d}"

def takip_basliklarini_hazirla(ws):
    eski_headers = baslik_haritasi(ws)
    if list(eski_headers.keys()) == TAKIP_KOLONLARI:
        return eski_headers

    eski_satirlar = []
    for row in range(2, ws.max_row + 1):
        kayit = {
            name: ws.cell(row=row, column=col).value
            for name, col in eski_headers.items()
        }
        if any(not hucre_bos_mu(value) for value in kayit.values()):
            eski_satirlar.append(kayit)

    ws.delete_rows(1, ws.max_row)
    ws.append(TAKIP_KOLONLARI)

    for idx, eski in enumerate(eski_satirlar, start=2):
        hisse = hisse_kodu_temizle(eski.get("Hisse") or eski.get("Hisse Adı"))
        al_tarihi = tarih_degeri(eski.get("AL Tarihi") or eski.get("Alış Tarihi"))
        cikis_tarihi = tarih_degeri(eski.get("Çıkış Tarihi") or eski.get("Satış Tarihi"))
        giris_fiyati = eski.get("Giriş Fiyatı")
        guncel_fiyat = eski.get("Güncel Fiyat")
        cikis_fiyati = eski.get("Çıkış Fiyatı")
        cikis_nedeni = eski.get("Çıkış Nedeni") or eski.get("KAR STOP/STOP")
        durum = eski.get("Durum") or (cikis_nedeni if not hucre_bos_mu(cikis_tarihi) else "AÇIK")
        stop_fiyati = eski.get("Stop Fiyatı")
        if hucre_bos_mu(stop_fiyati) and not hucre_bos_mu(giris_fiyati):
            stop_fiyati = yuvarla(float(giris_fiyati) * (1.0 - STOP_LOSS_PCT / 100.0))

        yeni = {col: None for col in TAKIP_KOLONLARI}
        yeni["İşlem ID"] = eski.get("İşlem ID") or takip_islem_id(idx)
        yeni["Hisse"] = hisse
        yeni["Sektör"] = eski.get("Sektör")
        yeni["AL Tarihi"] = al_tarihi
        yeni["AL Gücü"] = eski.get("AL Gücü")
        yeni["Giriş Fiyatı"] = giris_fiyati
        yeni["Stop Fiyatı"] = stop_fiyati
        yeni["Durum"] = durum
        yeni["Çıkış Tarihi"] = cikis_tarihi
        yeni["Çıkış Fiyatı"] = cikis_fiyati
        yeni["Çıkış Nedeni"] = cikis_nedeni
        yeni["Güncel Fiyat"] = guncel_fiyat
        yeni["Gerçekleşen Getiri (%)"] = getiri_yuzde(giris_fiyati, cikis_fiyati)
        yeni["Güncel Getiri (%)"] = None if not hucre_bos_mu(cikis_tarihi) else getiri_yuzde(giris_fiyati, guncel_fiyat)
        yeni["Pozisyonda Gün"] = (cikis_tarihi or datetime.now().date()) - al_tarihi if al_tarihi else None
        if yeni["Pozisyonda Gün"] is not None:
            yeni["Pozisyonda Gün"] = yeni["Pozisyonda Gün"].days
        yeni["Not"] = eski.get("Not")
        ws.append([yeni[col] for col in TAKIP_KOLONLARI])

    return baslik_haritasi(ws)

def acik_pozisyon_mu(ws, row, headers):
    durum = str(hucre_degeri(ws, row, headers, "Durum") or "").strip().upper()
    if durum == "AÇIK":
        return True
    cikis_tarihi = hucre_degeri(ws, row, headers, "Çıkış Tarihi", "Satış Tarihi")
    cikis_fiyati = hucre_degeri(ws, row, headers, "Çıkış Fiyatı")
    return hucre_bos_mu(cikis_tarihi) and hucre_bos_mu(cikis_fiyati)

def kapali_satir_metrikleri_tamam(ws, row, headers):
    if acik_pozisyon_mu(ws, row, headers):
        return False

    zorunlu_kolonlar = [
        "Çıkış Tarihi",
        "Çıkış Fiyatı",
        "Çıkış Nedeni",
        "Gerçekleşen Getiri (%)",
        "Pozisyonda Gün",
        "En Yüksek Fiyat",
        "En Düşük Fiyat",
        "MFE (%)",
        "MAE (%)",
        "BIST Giriş",
        "BIST Çıkış",
        "BIST Getiri (%)",
        "Alpha (%)",
    ]
    return all(
        not hucre_bos_mu(ws.cell(row=row, column=headers[col]).value)
        for col in zorunlu_kolonlar
        if col in headers
    )

def sozluk_degeri(data, *keys):
    for key in keys:
        if key in data:
            return data[key]
    return None

def bist_kapanis_getir(ws_bist, hedef_tarih):
    if ws_bist is None or hedef_tarih is None:
        return None

    headers = baslik_haritasi(ws_bist)
    tarih_col = headers.get("Tarih")
    kapanis_col = headers.get("Kapanış")
    if tarih_col is None or kapanis_col is None:
        return None

    aday = None
    for row in range(2, ws_bist.max_row + 1):
        tarih = tarih_degeri(ws_bist.cell(row=row, column=tarih_col).value)
        if tarih is None or tarih > hedef_tarih:
            continue
        aday = ws_bist.cell(row=row, column=kapanis_col).value
    return yuvarla(aday) if not hucre_bos_mu(aday) else None

def canli_fiyat_metrikleri_hesapla(df, alis_tarihi, giris_fiyati):
    if alis_tarihi is None or giris_fiyati is None or giris_fiyati <= 0:
        return None

    veri = df.copy()
    veri.index = pd.to_datetime(veri.index).date
    guncel_veri = veri[veri.index > alis_tarihi]
    if guncel_veri.empty:
        guncel_veri = veri[veri.index >= alis_tarihi]
    if guncel_veri.empty:
        return None

    en_yuksek = float(guncel_veri["High"].max())
    en_dusuk = float(guncel_veri["Low"].min())
    guncel_fiyat = float(guncel_veri["Close"].iloc[-1])

    return {
        "guncel": yuvarla(guncel_fiyat),
        "en_yuksek": yuvarla(en_yuksek),
        "en_dusuk": yuvarla(en_dusuk),
        "mfe": getiri_yuzde(giris_fiyati, en_yuksek),
        "mae": getiri_yuzde(giris_fiyati, en_dusuk),
    }

def pozisyon_cikis_hesapla(
    df,
    alis_tarihi,
    giris_fiyati,
    stop_fiyati=None,
    mevcut_cikis_tarihi=None,
    mevcut_cikis_fiyati=None,
    mevcut_cikis_nedeni=None,
):
    if alis_tarihi is None or giris_fiyati is None or giris_fiyati <= 0:
        return None

    veri = df.copy()
    veri.index = pd.to_datetime(veri.index).date
    guncel_veri = veri[veri.index >= alis_tarihi]
    if guncel_veri.empty:
        return None

    stop_fiyat = float(stop_fiyati) if not hucre_bos_mu(stop_fiyati) else giris_fiyati * (1.0 - STOP_LOSS_PCT / 100.0)
    zirve = giris_fiyati
    dip = giris_fiyati
    cikis_tarihi = mevcut_cikis_tarihi
    cikis_fiyati = float(mevcut_cikis_fiyati) if not hucre_bos_mu(mevcut_cikis_fiyati) else None
    cikis_nedeni = str(mevcut_cikis_nedeni).strip().upper() if not hucre_bos_mu(mevcut_cikis_nedeni) else None
    elle_kapali = cikis_tarihi is not None or cikis_fiyati is not None or cikis_nedeni in ("STOP", "KAR STOP")

    if elle_kapali and cikis_tarihi is None:
        cikis_tarihi = guncel_veri.index[-1]

    kontrol_veri = guncel_veri[guncel_veri.index > alis_tarihi]
    if elle_kapali and cikis_tarihi is not None:
        kontrol_veri = kontrol_veri[kontrol_veri.index <= cikis_tarihi]

    for idx, row in kontrol_veri.iterrows():
        high_i = float(row["High"])
        low_i = float(row["Low"])
        close_i = float(row["Close"])
        zirve = max(zirve, high_i)
        dip = min(dip, low_i)

        if elle_kapali:
            continue

        if low_i <= stop_fiyat:
            cikis_tarihi = idx
            cikis_fiyati = round(stop_fiyat, 2)
            cikis_nedeni = "STOP"
            break

        zirve_kar_pct = (zirve / giris_fiyati - 1.0) * 100.0
        geri_cekilme_pct = (zirve / close_i - 1.0) * 100.0 if close_i > 0 else 0.0
        if zirve_kar_pct >= PROFIT_TRIGGER_PCT and geri_cekilme_pct >= PULLBACK_PCT:
            cikis_tarihi = idx
            cikis_fiyati = round(close_i, 2)
            cikis_nedeni = "KAR STOP"
            break

    son_close = round(float(guncel_veri["Close"].iloc[-1]), 2)
    if len(kontrol_veri) == 0:
        zirve = max(zirve, float(guncel_veri["High"].iloc[-1]))
        dip = min(dip, float(guncel_veri["Low"].iloc[-1]))

    bitis_tarihi = cikis_tarihi or guncel_veri.index[-1]
    if cikis_nedeni is None and cikis_fiyati is not None:
        cikis_nedeni = "STOP" if cikis_fiyati <= stop_fiyat else "KAR STOP"

    return {
        "tarih": cikis_tarihi,
        "fiyat": yuvarla(cikis_fiyati),
        "neden": cikis_nedeni,
        "durum": cikis_nedeni or "AÇIK",
        "guncel": son_close,
        "gerceklesen_getiri": getiri_yuzde(giris_fiyati, cikis_fiyati) if cikis_fiyati is not None else None,
        "guncel_getiri": None if cikis_fiyati is not None else getiri_yuzde(giris_fiyati, son_close),
        "pozisyonda_gun": (bitis_tarihi - alis_tarihi).days,
        "en_yuksek": yuvarla(zirve),
        "en_dusuk": yuvarla(dip),
        "mfe": getiri_yuzde(giris_fiyati, zirve),
        "mae": getiri_yuzde(giris_fiyati, dip),
        "stop_sonrasi_guncel": getiri_yuzde(cikis_fiyati, son_close) if cikis_nedeni == "STOP" else None,
    }

def takip_excel_guncelle(al_listesi):
    if not TAKIP_EXCEL_GUNCELLE:
        return
    if load_workbook is None:
        print("  Takip Excel guncellenemedi: openpyxl kurulu degil.")
        return

    dosya = takip_dosyasi_bul()
    if not dosya.exists():
        print(f"  Takip Excel bulunamadi: {dosya}")
        return

    wb = load_workbook(dosya)
    if TAKIP_EXCEL_SAYFA not in wb.sheetnames:
        print(f"  Takip Excel guncellenemedi: '{TAKIP_EXCEL_SAYFA}' sayfasi yok.")
        wb.close()
        return

    ws = wb[TAKIP_EXCEL_SAYFA]
    ws_bist = wb["BIST"] if "BIST" in wb.sheetnames else None
    headers = takip_basliklarini_hazirla(ws)
    gerekli = TAKIP_KOLONLARI
    eksik = [col for col in gerekli if col not in headers]
    if eksik:
        print(f"  Takip Excel guncellenemedi: eksik kolon(lar): {', '.join(eksik)}")
        wb.close()
        return

    fiyat_cache = {}
    guncellenen = 0
    kapanan = 0

    for row in range(2, ws.max_row + 1):
        hisse = hisse_kodu_temizle(ws.cell(row=row, column=headers["Hisse"]).value)
        if not hisse:
            continue

        try:
            if hisse not in fiyat_cache:
                df, veri_kaynagi = veri_cek_kaynakli(hisse + ".IS", PERIOD_1D, INTERVAL)
                fiyat_cache[hisse] = (df, veri_kaynagi)
            df, _ = fiyat_cache[hisse]
            if df is None or df.empty:
                continue

            alis_tarihi = tarih_degeri(ws.cell(row=row, column=headers["AL Tarihi"]).value)
            giris_fiyati = ws.cell(row=row, column=headers["Giriş Fiyatı"]).value
            giris_fiyati = float(giris_fiyati) if not hucre_bos_mu(giris_fiyati) else None
            once_acik = acik_pozisyon_mu(ws, row, headers)
            sadece_canli_guncelle = kapali_satir_metrikleri_tamam(ws, row, headers)
            if sadece_canli_guncelle:
                canli = canli_fiyat_metrikleri_hesapla(df, alis_tarihi, giris_fiyati)
                if canli is None:
                    continue

                ws.cell(row=row, column=headers["Güncel Fiyat"]).value = canli["guncel"]
                ws.cell(row=row, column=headers["En Yüksek Fiyat"]).value = canli["en_yuksek"]
                ws.cell(row=row, column=headers["En Düşük Fiyat"]).value = canli["en_dusuk"]
                ws.cell(row=row, column=headers["MFE (%)"]).value = canli["mfe"]
                ws.cell(row=row, column=headers["MAE (%)"]).value = canli["mae"]

                cikis_nedeni = str(ws.cell(row=row, column=headers["Çıkış Nedeni"]).value or "").strip().upper()
                cikis_fiyati = ws.cell(row=row, column=headers["Çıkış Fiyatı"]).value
                if cikis_nedeni == "STOP" and not hucre_bos_mu(cikis_fiyati):
                    ws.cell(row=row, column=headers["Stop Sonrası Güncel (%)"]).value = getiri_yuzde(cikis_fiyati, canli["guncel"])
                guncellenen += 1
                continue

            stop_fiyati = ws.cell(row=row, column=headers["Stop Fiyatı"]).value
            if hucre_bos_mu(stop_fiyati) and giris_fiyati is not None:
                stop_fiyati = round(giris_fiyati * (1.0 - STOP_LOSS_PCT / 100.0), 2)
                ws.cell(row=row, column=headers["Stop Fiyatı"]).value = stop_fiyati

            mevcut_cikis_tarihi = tarih_degeri(ws.cell(row=row, column=headers["Çıkış Tarihi"]).value)
            mevcut_cikis_fiyati = ws.cell(row=row, column=headers["Çıkış Fiyatı"]).value
            mevcut_cikis_nedeni = ws.cell(row=row, column=headers["Çıkış Nedeni"]).value
            cikis = pozisyon_cikis_hesapla(
                df,
                alis_tarihi,
                giris_fiyati,
                stop_fiyati,
                mevcut_cikis_tarihi,
                mevcut_cikis_fiyati,
                mevcut_cikis_nedeni,
            )
            if cikis is None:
                continue

            bist_giris = ws.cell(row=row, column=headers["BIST Giriş"]).value
            if hucre_bos_mu(bist_giris):
                bist_giris = bist_kapanis_getir(ws_bist, alis_tarihi)
                ws.cell(row=row, column=headers["BIST Giriş"]).value = bist_giris

            bist_bitis_tarihi = cikis["tarih"] or pd.to_datetime(df.index[-1]).date()
            bist_cikis = bist_kapanis_getir(ws_bist, bist_bitis_tarihi)
            bist_getiri = getiri_yuzde(bist_giris, bist_cikis)
            hisse_getiri = cikis["gerceklesen_getiri"] if cikis["gerceklesen_getiri"] is not None else cikis["guncel_getiri"]

            ws.cell(row=row, column=headers["Güncel Fiyat"]).value = cikis["guncel"]
            ws.cell(row=row, column=headers["Durum"]).value = cikis["durum"]
            ws.cell(row=row, column=headers["Gerçekleşen Getiri (%)"]).value = cikis["gerceklesen_getiri"]
            ws.cell(row=row, column=headers["Güncel Getiri (%)"]).value = cikis["guncel_getiri"]
            ws.cell(row=row, column=headers["Pozisyonda Gün"]).value = cikis["pozisyonda_gun"]
            ws.cell(row=row, column=headers["En Yüksek Fiyat"]).value = cikis["en_yuksek"]
            ws.cell(row=row, column=headers["En Düşük Fiyat"]).value = cikis["en_dusuk"]
            ws.cell(row=row, column=headers["MFE (%)"]).value = cikis["mfe"]
            ws.cell(row=row, column=headers["MAE (%)"]).value = cikis["mae"]
            ws.cell(row=row, column=headers["BIST Çıkış"]).value = bist_cikis
            ws.cell(row=row, column=headers["BIST Getiri (%)"]).value = bist_getiri
            ws.cell(row=row, column=headers["Alpha (%)"]).value = round(hisse_getiri - bist_getiri, 2) if hisse_getiri is not None and bist_getiri is not None else None
            ws.cell(row=row, column=headers["Stop Sonrası Güncel (%)"]).value = cikis["stop_sonrasi_guncel"]
            guncellenen += 1

            if cikis["tarih"] is not None:
                ws.cell(row=row, column=headers["Çıkış Tarihi"]).value = cikis["tarih"]
                ws.cell(row=row, column=headers["Çıkış Fiyatı"]).value = cikis["fiyat"]
                ws.cell(row=row, column=headers["Çıkış Nedeni"]).value = cikis["neden"]
                if once_acik:
                    kapanan += 1
        except Exception as e:
            print(f"  Takip Excel: {hisse} guncellenemedi - {e}")

    mevcut_acik_hisseler = set()
    mevcut_alis_kayitlari = set()
    for row in range(2, ws.max_row + 1):
        hisse = hisse_kodu_temizle(ws.cell(row=row, column=headers["Hisse"]).value)
        alis = tarih_degeri(ws.cell(row=row, column=headers["AL Tarihi"]).value)
        if hisse:
            mevcut_alis_kayitlari.add((hisse, alis))
        if hisse and acik_pozisyon_mu(ws, row, headers):
            mevcut_acik_hisseler.add(hisse)

    eklenen = 0
    for sinyal in al_listesi:
        hisse = hisse_kodu_temizle(sinyal.get("Hisse"))
        sinyal_tarihi = tarih_degeri(sinyal.get("Sinyal Tarihi")) or datetime.now().date()
        if not hisse or hisse in mevcut_acik_hisseler or (hisse, sinyal_tarihi) in mevcut_alis_kayitlari:
            continue

        yeni_row = ws.max_row + 1
        kapanis_fiyati = sozluk_degeri(sinyal, "Kapanış Fiyatı", "Kapanis Fiyati", "Kapanis_Fiyati")
        if kapanis_fiyati is None:
            continue

        giris_fiyati = float(kapanis_fiyati)
        stop_fiyati = sozluk_degeri(sinyal, "Stop Fiyatı", "Stop Fiyati", "Stop_Fiyati")
        stop_fiyati = float(stop_fiyati) if stop_fiyati is not None else round(giris_fiyati * (1.0 - STOP_LOSS_PCT / 100.0), 2)
        bist_giris = bist_kapanis_getir(ws_bist, sinyal_tarihi)

        ws.cell(row=yeni_row, column=headers["İşlem ID"]).value = takip_islem_id(yeni_row)
        ws.cell(row=yeni_row, column=headers["Hisse"]).value = hisse
        ws.cell(row=yeni_row, column=headers["Sektör"]).value = None
        ws.cell(row=yeni_row, column=headers["AL Tarihi"]).value = sinyal_tarihi
        ws.cell(row=yeni_row, column=headers["AL Gücü"]).value = sozluk_degeri(sinyal, "AL Gücü", "AL Gucu")
        ws.cell(row=yeni_row, column=headers["Giriş Fiyatı"]).value = giris_fiyati
        ws.cell(row=yeni_row, column=headers["Stop Fiyatı"]).value = stop_fiyati
        ws.cell(row=yeni_row, column=headers["Durum"]).value = "AÇIK"
        ws.cell(row=yeni_row, column=headers["Çıkış Tarihi"]).value = None
        ws.cell(row=yeni_row, column=headers["Çıkış Fiyatı"]).value = None
        ws.cell(row=yeni_row, column=headers["Çıkış Nedeni"]).value = None
        ws.cell(row=yeni_row, column=headers["Güncel Fiyat"]).value = giris_fiyati
        ws.cell(row=yeni_row, column=headers["Gerçekleşen Getiri (%)"]).value = None
        ws.cell(row=yeni_row, column=headers["Güncel Getiri (%)"]).value = 0.0
        ws.cell(row=yeni_row, column=headers["Pozisyonda Gün"]).value = 0
        ws.cell(row=yeni_row, column=headers["En Yüksek Fiyat"]).value = giris_fiyati
        ws.cell(row=yeni_row, column=headers["En Düşük Fiyat"]).value = giris_fiyati
        ws.cell(row=yeni_row, column=headers["MFE (%)"]).value = 0.0
        ws.cell(row=yeni_row, column=headers["MAE (%)"]).value = 0.0
        ws.cell(row=yeni_row, column=headers["BIST Giriş"]).value = bist_giris
        ws.cell(row=yeni_row, column=headers["BIST Çıkış"]).value = None
        ws.cell(row=yeni_row, column=headers["BIST Getiri (%)"]).value = None
        ws.cell(row=yeni_row, column=headers["Alpha (%)"]).value = None
        ws.cell(row=yeni_row, column=headers["Stop Sonrası Güncel (%)"]).value = None
        ws.cell(row=yeni_row, column=headers["Not"]).value = sozluk_degeri(sinyal, "Not") or "Ertesi gün açılışta giriş"
        mevcut_acik_hisseler.add(hisse)
        mevcut_alis_kayitlari.add((hisse, sinyal_tarihi))
        eklenen += 1

    wb.save(dosya)
    wb.close()
    print(f"  Takip Excel guncellendi: {dosya} | yeni AL: {eklenen}, guncel fiyat: {guncellenen}, kapanan: {kapanan}")

# ─────────────────────────────────────────────
# ANA TARAMA
# ─────────────────────────────────────────────
def tara():
    print("\n" + "="*60)
    print("  BIST ESv1 TARAMA - GUNLUK PERIYOT")
    print(f"  Tarih  : {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print(f"  Veri   : {DATA_SOURCE} | fallback: {'ACIK' if ALLOW_DATA_FALLBACK else 'KAPALI'}")
    print(f"  HTF    : {'ACIK' if USE_HTF else 'KAPALI'}")
    print(f"  Hacim  : {'ACIK' if USE_SIGNAL_VOLUME_FILTER else 'KAPALI'} | son mum >= {SIGNAL_VOLUME_MULTIPLIER}x ort")
    print(f"  MA20   : {'ACIK' if USE_TREND else 'KAPALI'} | {MA_SLOPE_BARS} bar >= %{MIN_MA_SLOPE_PCT}")
    print(f"  Hisse  : {len(BIST_HISSELER)} adet")
    print("  Son kapanan gunluk mumda sinyal araniyor")
    print("     -> Ertesi gun acilista giris yapilabilir")
    print("="*60)

    if DATA_SOURCE == "tradingview":
        ok, mesaj = tradingview_on_kontrol()
        print(f"  Veri on kontrol: {mesaj}")
        if not ok:
            print("  Tarama durduruldu. TradingView baglantisi kurulmadan sonuc tutarli olmaz.")
            print("  Cozum: Internet baglantisini kontrol edin veya .env dosyasina TV_USERNAME / TV_PASSWORD ekleyin.")
            print("="*60 + "\n")
            return

    al_listesi   = []
    hata_listesi = []
    toplam       = len(BIST_HISSELER)

    for idx, hisse in enumerate(BIST_HISSELER, 1):
        ticker = hisse + ".IS"
        print(f"  [{idx:3d}/{toplam}] {hisse:<10}", end=" ", flush=True)
        try:
            if USE_HTF and not htf_ok(ticker):
                print("- HTF engelledi")
                continue

            df, veri_kaynagi = veri_cek_kaynakli(ticker, PERIOD_1D, INTERVAL)
            if df is None or len(df) < max(30, MA_TREND_LEN + MA_SLOPE_BARS + 5, VOL_LEN + 5):
                hata = son_veri_kaynagi_hatasi()
                kaynak_text = veri_kaynagi if veri_kaynagi else "yok"
                print(f"! Veri yok | Veri: {kaynak_text}" + (f" | {hata}" if hata else ""))
                hata_listesi.append(hisse)
                continue

            al, sat, close, grade, stop_fiyat, sat_neden = sinyal_hesapla(df)

            if len(al) < 3:
                print("- Yetersiz veri")
                continue

            # Son kapanan mum (iloc[-1]) — günlük periyotta mum kapanmış olur
            # Tarama akşam yapıldığı için son mum kesinleşmiş
            son_al = bool(al.iloc[-1])
            son_sat = bool(sat.iloc[-1])

            if son_al:
                sinyal_tarihi = df.index[-1].strftime("%d.%m.%Y")
                sinyal_fiyat  = round(float(close.iloc[-1]), 2)
                stop_seviye    = round(float(stop_fiyat.iloc[-1]), 2)
                al_gucu        = buy_grade_text(int(grade.iloc[-1]))

                al_listesi.append({
                    "Hisse"          : hisse,
                    "Kapanış Fiyatı" : sinyal_fiyat,
                    "Stop Fiyatı"    : stop_seviye,
                    "AL Gücü"        : al_gucu,
                    "Sinyal Tarihi"  : sinyal_tarihi,
                    "Veri Kaynagi"   : veri_kaynagi,
                    "Not"            : "Ertesi gün açılışta giriş"
                })
                print(f"OK {al_gucu} - {sinyal_fiyat} TL | Stop {stop_seviye} TL  ({sinyal_tarihi}) | Veri: {veri_kaynagi}")
            elif son_sat:
                neden = sat_neden.iloc[-1] if sat_neden.iloc[-1] else "SAT"
                print(f"-- {neden} | Veri: {veri_kaynagi}")
            else:
                print(f"- Sinyal yok | Veri: {veri_kaynagi}")

        except Exception as e:
            print(f"X Hata: {e}")
            hata_listesi.append(hisse)

    # ─────────────────────────────────────────────
    # SONUÇLAR
    # ─────────────────────────────────────────────
    print("\n" + "="*60)
    print(f"  AL SINYALI VEREN HISSELER ({len(al_listesi)} adet)")
    print(f"  -> Yarin acilista giris yapilabilir")
    print("="*60)
    if al_listesi:
        for h in al_listesi:
            print(f"  {h['Hisse']:<10} {h['Kapanış Fiyatı']:>10.2f} TL   Stop: {h['Stop Fiyatı']:>10.2f} TL   {h['AL Gücü']:<10} {h['Sinyal Tarihi']}   Veri: {h['Veri Kaynagi']}")
    else:
        print("  Sinyal veren hisse bulunamadi.")

    if al_listesi and GUNLUK_AL_EXCEL_KAYDET:
        dosya = f"bist_al_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        pd.DataFrame(al_listesi).to_excel(dosya, index=False)
        print(f"\n  Excel kaydedildi: {dosya}")

    takip_excel_guncelle(al_listesi)

    if hata_listesi:
        print(f"\n  Veri alinamayan: {len(hata_listesi)} hisse")

    print("="*60 + "\n")
    try:
        input("  Cikmak icin Enter'a basin...")
    except EOFError:
        pass

if __name__ == "__main__":
    print("Gunluk tarama baslatiliyor...", flush=True)
    tara()
